import os
# Check if file exists and what sheets it has
if os.path.exists('leave_data.xlsx'):
    from openpyxl import load_workbook
    wb = load_workbook('leave_data.xlsx')
    print("Sheets found in file:", wb.sheetnames)
else:
    print("File does not exist yet.")