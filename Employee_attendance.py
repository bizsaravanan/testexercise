import PySimpleGUI as sg
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

EXCEL_FILE = 'leave_data.xlsx'


# --- Initialization ---
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws_emp = wb.active
        ws_emp.title = "Employees"
        ws_emp.append(['ID', 'Name', 'CL', 'SL', 'PL'])
        ws_log = wb.create_sheet("Logs")
        ws_log.append(['ID', 'Name', 'Date', 'Type', 'Reason'])
        wb.save(EXCEL_FILE)
    else:
        wb = load_workbook(EXCEL_FILE)
        if 'Logs' not in wb.sheetnames:
            ws_log = wb.create_sheet("Logs")
            ws_log.append(['ID', 'Name', 'Date', 'Type', 'Reason'])
            wb.save(EXCEL_FILE)


def get_employees():
    if not os.path.exists(EXCEL_FILE): return []
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name='Employees', engine='openpyxl', dtype={'ID': str})
        return [f"{row['ID']} - {row['Name']}" for _, row in df.iterrows()]
    except:
        return []


initialize_excel()

# --- UI Layout ---
layout = [
    [sg.Frame("Register Employee", [
        [sg.Text("ID:"), sg.Input(key="-ID-", size=(10, 1)), sg.Text("Name:"), sg.Input(key="-NAME-", size=(20, 1))],
        [sg.Text("CL:"), sg.Input(key="-CL-", size=(5, 1)), sg.Text("SL:"), sg.Input(key="-SL-", size=(5, 1)),
         sg.Text("PL:"), sg.Input(key="-PL-", size=(5, 1))],
        [sg.Button("Save Employee")]
    ])],
    [sg.Frame("Apply for Leave", [
        [sg.Text("Select Employee:"),
         sg.Combo(get_employees(), key="-EMP-", enable_events=True, readonly=True, size=(30, 1))],
        [sg.Text("Balance:", key="-BAL-", text_color="yellow")],
        [sg.Text("Type:"), sg.Combo(["CL", "SL", "PL"], key="-TYPE-", readonly=True)],
        [sg.Text("From:"), sg.Input(key="-FROM-", size=(10, 1)),
         sg.CalendarButton("Cal", target="-FROM-", format='%Y-%m-%d'),
         sg.Text("To:"), sg.Input(key="-TO-", size=(10, 1)),
         sg.CalendarButton("Cal", target="-TO-", format='%Y-%m-%d')],
        [sg.Text("Reason:"), sg.Input(key="-REASON-", size=(30, 1)), sg.Button("Apply")]
    ])],
    [sg.Table(values=[], headings=['ID', 'Name', 'Date', 'Type', 'Reason'],
              auto_size_columns=False, justification='center',
              col_widths=[5, 20, 20, 4, 15],
              key="-TABLE-", num_rows=10)]
]

window = sg.Window("Leave Management System", layout)

# --- Event Loop ---
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED: break

    if event == "Save Employee":
        emp_df = pd.read_excel(EXCEL_FILE, sheet_name='Employees', engine='openpyxl', dtype={'ID': str})
        new_emp = {
            'ID': str(values['-ID-']), 'Name': values['-NAME-'],
            'CL': int(values['-CL-']), 'SL': int(values['-SL-']), 'PL': int(values['-PL-'])
        }
        emp_df = pd.concat([emp_df, pd.DataFrame([new_emp])], ignore_index=True)
        log_df = pd.read_excel(EXCEL_FILE, sheet_name='Logs', engine='openpyxl', dtype={'ID': str})

        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            emp_df.to_excel(writer, sheet_name='Employees', index=False)
            log_df.to_excel(writer, sheet_name='Logs', index=False)
        window["-EMP-"].update(values=get_employees())
        sg.popup("Employee Saved!")

    if event == "-EMP-":
        eid = values["-EMP-"].split(" - ")[0]
        emp_df = pd.read_excel(EXCEL_FILE, sheet_name='Employees', engine='openpyxl', dtype={'ID': str})
        row = emp_df[emp_df['ID'] == eid].iloc[0]
        window["-BAL-"].update(f"CL: {row['CL']} | SL: {row['SL']} | PL: {row['PL']}")
        log_df = pd.read_excel(EXCEL_FILE, sheet_name='Logs', engine='openpyxl', dtype={'ID': str})
        window["-TABLE-"].update(values=log_df[log_df['ID'] == eid].values.tolist())

    if event == "Apply":
        if not values["-EMP-"] or not values['-FROM-'] or not values['-TO-']:
            sg.popup_error("Please fill all fields!")
            continue

        # Calculate duration
        d1 = datetime.strptime(values['-FROM-'], '%Y-%m-%d')
        d2 = datetime.strptime(values['-TO-'], '%Y-%m-%d')
        duration = (d2 - d1).days + 1

        eid = values["-EMP-"].split(" - ")[0]
        l_type = values['-TYPE-']

        emp_df = pd.read_excel(EXCEL_FILE, sheet_name='Employees', engine='openpyxl', dtype={'ID': str})
        log_df = pd.read_excel(EXCEL_FILE, sheet_name='Logs', engine='openpyxl', dtype={'ID': str})

        # Deduct balance
        idx = emp_df[emp_df['ID'] == eid].index[0]
        if emp_df.at[idx, l_type] >= duration:
            emp_df.at[idx, l_type] -= duration
            new_log = {'ID': eid, 'Name': values["-EMP-"].split(" - ")[1],
                       'Date': f"{values['-FROM-']} to {values['-TO-']}", 'Type': l_type, 'Reason': values['-REASON-']}
            log_df = pd.concat([log_df, pd.DataFrame([new_log])], ignore_index=True)

            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                emp_df.to_excel(writer, sheet_name='Employees', index=False)
                log_df.to_excel(writer, sheet_name='Logs', index=False)

            window["-BAL-"].update(
                f"CL: {emp_df.at[idx, 'CL']} | SL: {emp_df.at[idx, 'SL']} | PL: {emp_df.at[idx, 'PL']}")
            window["-TABLE-"].update(values=log_df[log_df['ID'] == eid].values.tolist())
            sg.popup("Leave Applied successfully!")
        else:
            sg.popup_error(f"Insufficient {l_type} balance!")

window.close()